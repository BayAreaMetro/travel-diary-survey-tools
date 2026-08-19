"""One-time extraction of the ETC vendor test workbook into stable CSVs.

The vendor shipped the test data as a single .xlsx.  Rather than make the
pipeline depend on an Excel reader for one file, this script is run once to
land plain CSVs that ``load_data`` reads like any other survey input.  Re-run
it only if the vendor sends a new workbook.

Survey data never lives in this repo: the CSVs are written to an
``extracted_data`` folder beside the source workbook on Box, and the pipeline
config points at them there.

The workbook also carries pivot caches that break openpyxl's default
(read/write) loader, so it is opened read-only here -- one more reason the
pipeline should never touch the .xlsx directly.

Usage:
    python projects/etc_test/extract_xlsx.py [--workbook PATH] [--out DIR]
"""

import argparse
import csv
import datetime as dt
import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

VENDOR_DIR = Path(
    "E:/Box/Modeling and Surveys/Surveys/Travel Diary Survey/BATS_2028"
    "/BATS_28_MTC_Partners_Shared/Pre-solicitation research/Vendor Responses"
    "/ETC_app_files"
)
WORKBOOK = VENDOR_DIR / "San Fransisco (MTC) HHTS - Test Data 1.xlsx"
OUT_DIR = VENDOR_DIR / "extracted_data"


@dataclass(frozen=True)
class SheetSpec:
    """A workbook sheet and how to land it as CSV.

    Attributes:
        sheet: Sheet name as it appears in the workbook.
        csv_name: Output file name.
        header_row: 1-based row holding the column names.
    """

    sheet: str
    csv_name: str
    header_row: int


# The four data tables plus the two documentation sheets.  The workbook's other
# sheets are not data and are deliberately skipped: "Pivots" summarises a
# larger dataset that is not in this file, "Shape Tagging" holds another
# study's points (Pima County, AZ), and "Sheet1" is a column-letter helper list.
SHEETS = (
    SheetSpec("1 Household", "households.csv", 1),
    SheetSpec("2 Person", "persons.csv", 1),
    SheetSpec("3 Vehicle", "vehicles.csv", 1),
    SheetSpec("4 Trips", "trips.csv", 1),
    SheetSpec("Data Dictionary", "data_dictionary.csv", 3),
    SheetSpec("Codebook", "codebook.csv", 3),
)


def _cell_to_str(value: object) -> str:
    """Render one cell as text that round-trips through CSV unambiguously.

    Excel hands back datetimes for date-only columns and floats for integer
    codes; both are normalised so the CSVs stay stable across re-extraction
    and so codes do not arrive as "20.0".
    """
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        # Date-only columns come back with a midnight time component.
        return value.date().isoformat() if value.time() == dt.time.min else value.isoformat(" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _keep_columns(header: list[str], rows: list[list[str]]) -> list[int]:
    """Return the column indices worth writing.

    Drops columns that are empty *and* unnamed -- the spacer columns to the
    left of a header that starts at column B.  A named column with no values is
    kept: it is a real vendor field that this small test extract happens not to
    populate.
    """
    width = max((len(r) for r in [header, *rows]), default=0)
    return [
        col
        for col in range(width)
        if (col < len(header) and header[col]) or any(col < len(r) and r[col] for r in rows)
    ]


def extract_sheet(workbook: object, spec: SheetSpec, out_dir: Path) -> tuple[int, int]:
    """Write one sheet to CSV.

    Args:
        workbook: An open, read-only openpyxl workbook.
        spec: Which sheet to read and where to put it.
        out_dir: Directory to write the CSV into.

    Returns:
        Tuple of (data rows written, columns written).
    """
    ws = workbook[spec.sheet]
    rows = [[_cell_to_str(v) for v in row] for row in ws.iter_rows(values_only=True)]

    # Excel reports a dimension far past the real data on some sheets; drop the
    # rows above the header and any row that is blank all the way across.
    header, *data = rows[spec.header_row - 1 :]
    data = [r for r in data if any(r)]

    keep = _keep_columns(header, data)
    trimmed = [[r[c] if c < len(r) else "" for c in keep] for r in [header, *data]]

    with (out_dir / spec.csv_name).open("w", newline="", encoding="utf-8") as fh:
        csv.writer(fh, lineterminator="\n").writerows(trimmed)

    return len(data), len(keep)


def main() -> None:
    """Extract every sheet of interest into CSVs."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK, help="Source .xlsx")
    parser.add_argument("--out", type=Path, default=OUT_DIR, help="Output directory for CSVs")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(message)s")
    args.out.mkdir(parents=True, exist_ok=True)

    # read_only avoids openpyxl's pivot-cache parsing, which raises on this file.
    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    try:
        logger.info("Extracting %s", args.workbook.name)
        for spec in SHEETS:
            n_rows, n_cols = extract_sheet(workbook, spec, args.out)
            logger.info(
                "  %-16s -> %-22s %4d rows x %3d cols", spec.sheet, spec.csv_name, n_rows, n_cols
            )
    finally:
        workbook.close()

    logger.info("Wrote %d CSVs to %s", len(SHEETS), args.out)


if __name__ == "__main__":
    main()
